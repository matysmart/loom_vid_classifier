import openpyxl

def ensure_video_header_and_shift(file_path):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Check if A1 contains "video", if not, shift the column down and add it
    if ws['A1'].value != 'video':
        # Shift the entire column A downward by 1 row
        ws.insert_rows(1)
        ws['A1'].value = 'video'
        print(f'Shifted column A down and added "video" header to A1.')
    else:
        print(f'"video" header already exists in A1.')

    # Save the workbook
    wb.save(file_path)

# Example usage
file_path = 'linksNoHeader.xlsx'
ensure_video_header_and_shift(file_path)
